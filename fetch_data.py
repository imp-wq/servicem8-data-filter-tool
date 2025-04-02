import http.server
import json
import logging
import os
import socketserver
import threading
import urllib.parse
import webbrowser
from openpyxl import Workbook, load_workbook

import jmespath
import requests
from requests_oauthlib import OAuth2Session

from config import SERVICEM8_CLIENT_ID, SERVICEM8_CLIENT_SECRET, SCOPE

client_id = SERVICEM8_CLIENT_ID
client_secret = SERVICEM8_CLIENT_SECRET
authorization_base_url = 'https://go.servicem8.com/oauth/authorize'
token_url = 'https://go.servicem8.com/oauth/token'
redirect_uri = 'http://localhost:8080/callback'

BASE_URL = "https://api.servicem8.com/api_1.0"

FIELD_KEY = "created_by_staff_uuid"
# STAFF_UUID = "71a2407f-2970-454a-acf8-229357b3d11b"

STAFF_FIRST = "Overflow Solutions PTY LTD"
STAFF_LAST = "(Partner Support)"

FILE_PATH = "jobs.xlsx"

DISPLAYED_KEY_LIST = [
    "generated_job_id",
    "date",
    "created_by_staff_uuid"
]
SHEET_TITLE = "Jobs"

logging.basicConfig(level=logging.INFO)


class OAuthHandler(http.server.SimpleHTTPRequestHandler):
    def do_GET(self):
        global auth_code
        # Parse query parameters from the URL.
        parsed_url = urllib.parse.urlparse(self.path)
        query_params = urllib.parse.parse_qs(parsed_url.query)
        if 'code' in query_params:
            auth_code = query_params['code'][0]

            # Send a simple response to the browser.
            self.send_response(200)
            self.send_header('Content-type', 'text/html')
            self.end_headers()
            self.wfile.write(
                b"<html><body><h1>Authentication successful. You may now close this window.</h1></body></html>")
        else:
            self.send_response(400)
            self.end_headers()


def run_local_server():
    with socketserver.TCPServer(("localhost", 8080), OAuthHandler) as httpd:
        # Handle a single request then shutdown.
        httpd.handle_request()


def run_oauth_flow():
    # Create an OAuth2 session.
    oauth = OAuth2Session(client_id, redirect_uri=redirect_uri, scope=" ".join(SCOPE))

    # Generate the authorization URL.
    authorization_url, state = oauth.authorization_url(authorization_base_url)

    # Automatically open the browser to the authorization URL.
    print(state)
    webbrowser.open(authorization_url)
    print("A browser window has been opened for authentication...")

    # Run the local server in a separate thread to capture the OAuth redirect.
    server_thread = threading.Thread(target=run_local_server)
    server_thread.start()
    server_thread.join()

    # Exchange the authorization code for an access token.
    token = oauth.fetch_token(token_url, client_secret=client_secret, code=auth_code)

    return token['access_token']


def query_staff_uuid(access_token: str):
    url = f"{BASE_URL}/staff.json"
    headers = {
        "Authorization": f"Bearer {access_token}",
    }
    response = requests.get(url, headers=headers)

    staff_list = response.json()
    search_by_name = jmespath.search(f"[?first=='{STAFF_FIRST}' && last=='{STAFF_LAST}']", staff_list)
    search_uuid = jmespath.search("[].uuid", search_by_name)
    if len(search_uuid) <= 0:
        logging.error(f"""No staff whose:
        first='{STAFF_FIRST}
        last={STAFF_LAST}'
        can be found""")
        raise ValueError()
    uuid = search_uuid[0]
    return uuid


def query_jobs(access_token: str, staff_uuid: str):
    url = f"{BASE_URL}/job.json"
    headers = {
        "Authorization": f"Bearer {access_token}",
    }
    response = requests.get(url, headers=headers)
    # print(response.json())
    jobs = response.json()

    # Use jmespath to filter for jobs where staff_uuid equals the given_uuid
    query = f"[?{FIELD_KEY}=='{staff_uuid}']"
    filtered_jobs = jmespath.search(query, jobs)

    # Dynamically create the projection string from DISPLAYED_KEY_LIST.
    projection = ", ".join([f"{field}: {field}" for field in DISPLAYED_KEY_LIST])

    # Build the jmespath query.
    displayed_jobs = jmespath.search(f"[].{{{projection}}}", filtered_jobs)

    return displayed_jobs


def write_jobs_to_excel(jobs: list, file_path: str = FILE_PATH, sheet_title: str = SHEET_TITLE):
    if not jobs:
        logging.error("No job data to write.")
        raise ValueError()

    headers = list(jobs[0].keys())

    if os.path.exists(file_path):
        logging.info(f"File {file_path} already exists, appending to existing file.")
        # File exists: open it
        workbook = load_workbook(file_path)

        # Check if the sheet exists, create if it doesn't
        if SHEET_TITLE in workbook.sheetnames:
            sheet = workbook[SHEET_TITLE]
        else:
            sheet = workbook.create_sheet(title=SHEET_TITLE)
            logging.info(f"Sheet '{SHEET_TITLE}' not found. Created a new one.")

            # Write headers in the new sheet
            for col_index, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_index, value=header)

        start_row = sheet.max_row + 1
    else:
        logging.info(f"File {file_path} does not exist, creating new file.")
        # File does not exist: create it
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title
        start_row = 2  # Start from second row, since we'll write headers
        # Write headers
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_index, value=header)

    # Write job data
    for i, job in enumerate(jobs):
        for j, key in enumerate(headers):
            sheet.cell(row=start_row + i, column=j + 1, value=job.get(key, ""))

    workbook.save(file_path)
    logging.info(f"File {file_path} written {len(jobs)} rows.")


def test(access_token: str):
    url = f"{BASE_URL}/account.json"
    headers = {
        "Authorization": f"Bearer {access_token}",
    }
    response = requests.get(url, headers=headers)
    print(response)


def main():
    access_token = run_oauth_flow()
    logging.info(access_token)
    test(access_token)
    # staff_uuid = query_staff_uuid(access_token)
    # jobs = query_jobs(access_token, staff_uuid)
    # write_jobs_to_excel(jobs)


if __name__ == '__main__':
    main()
